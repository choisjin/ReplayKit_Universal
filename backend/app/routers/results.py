"""Test results API routes."""

import asyncio
import base64
import gzip
import html as _html
import io
import json
import logging
import os
import re
import shutil
import subprocess
import sys          # sys.platform — open_folder(1270행)/trim_recording 의 Windows 분기
import tempfile
import threading
import time
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import FileResponse, RedirectResponse, Response, StreamingResponse
from starlette.background import BackgroundTask

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/results", tags=["results"])

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent  # server.py 위치
RESULTS_DIR = Path(__file__).resolve().parent.parent.parent / "results"
SCREENSHOTS_DIR = Path(__file__).resolve().parent.parent.parent / "screenshots"
RECORDINGS_DIR = _PROJECT_ROOT / "Results" / "Video"
EXPORT_ROOT = _PROJECT_ROOT / "Results"
_TOOLS_DIR = _PROJECT_ROOT / "tools"


def _find_ffmpeg() -> str | None:
    """ffmpeg 실행 파일 경로를 반환. 시스템 PATH → tools/ 폴더 순으로 탐색."""
    # 시스템 PATH
    found = shutil.which("ffmpeg")
    if found:
        return found
    # 프로젝트 tools/ 폴더
    local = _TOOLS_DIR / "ffmpeg.exe"
    if local.is_file():
        return str(local)
    # tools/ffmpeg/bin/ 구조 (일반적인 ffmpeg 배포 패키지)
    local_bin = _TOOLS_DIR / "ffmpeg" / "bin" / "ffmpeg.exe"
    if local_bin.is_file():
        return str(local_bin)
    return None


def _content_disposition(filename: str) -> str:
    """RFC 5987 형식의 Content-Disposition 헤더 값 생성.

    HTTP 헤더 값은 latin-1로만 인코딩되므로 한글·괄호 등 비-ASCII 파일명을
    그대로 넣으면 UnicodeEncodeError로 응답이 깨진다. ASCII fallback(filename=)과
    UTF-8 인코딩(filename*=)을 함께 제공해 한글 파일명도 안전하게 내려보낸다.
    """
    ascii_fallback = filename.encode("ascii", "ignore").decode().strip() or "result"
    return (
        f'attachment; filename="{ascii_fallback}"; '
        f"filename*=UTF-8''{quote(filename)}"
    )


@router.get("/list")
async def list_results():
    """List all test result files (런 폴더 + 레거시 플랫 파일 모두 탐색).

    결과가 쌓이면 전체 result.json을 요약용으로 파싱하는 비용이 커서
    이벤트 루프를 막는다 → 스레드로 오프로드.
    """
    return await asyncio.to_thread(_list_results_sync)


def _list_results_sync():
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    results = []
    seen: set[str] = set()

    # 1) 런 폴더: results/{ts}_{scenario}/result.json
    for d in sorted(RESULTS_DIR.iterdir(), reverse=True):
        if not d.is_dir():
            continue
        rj = d / "result.json"
        if not rj.exists():
            continue
        try:
            data = json.loads(rj.read_text(encoding="utf-8"))
        except Exception:
            continue
        key = d.name
        seen.add(key)
        results.append({
            "filename": f"{d.name}/result.json",
            "run_folder": d.name,
            "scenario_name": data.get("scenario_name", ""),
            "status": data.get("status", ""),
            "total_steps": data.get("total_steps", 0),
            "total_repeat": data.get("total_repeat", 1),
            "passed_steps": data.get("passed_steps", 0),
            "failed_steps": data.get("failed_steps", 0),
            "warning_steps": data.get("warning_steps", 0),
            "error_steps": data.get("error_steps", 0),
            "started_at": data.get("started_at", ""),
            "finished_at": data.get("finished_at", ""),
        })

    # 2) 레거시 플랫: results/*.json
    for f in sorted(RESULTS_DIR.glob("*.json"), reverse=True):
        data = json.loads(f.read_text(encoding="utf-8"))
        results.append({
            "filename": f.name,
            "run_folder": "",
            "scenario_name": data.get("scenario_name", ""),
            "status": data.get("status", ""),
            "total_steps": data.get("total_steps", 0),
            "total_repeat": data.get("total_repeat", 1),
            "passed_steps": data.get("passed_steps", 0),
            "failed_steps": data.get("failed_steps", 0),
            "warning_steps": data.get("warning_steps", 0),
            "error_steps": data.get("error_steps", 0),
            "started_at": data.get("started_at", ""),
            "finished_at": data.get("finished_at", ""),
        })
    return {"results": results}


def _resolve_image_path(rel_path: str | None) -> Path | None:
    """Resolve a relative screenshot path to an absolute filesystem path.

    actual 이미지는 run 폴더(RESULTS_DIR/{run}/screenshots/...) 기준으로 저장되고,
    expected 이미지는 SCREENSHOTS_DIR/{scenario}/... 기준으로 저장된다.
    둘 다 시도한다.
    """
    if not rel_path:
        return None
    p = rel_path.replace("\\", "/")
    # 절대 경로면 그대로 시도
    try:
        ap = Path(p)
        if ap.is_absolute() and ap.exists():
            return ap
    except Exception:
        pass
    # 옛 결과의 .../screenshots/ 접두사 제거 (SCREENSHOTS_DIR 기준 잔존 케이스용)
    idx = p.find("/screenshots/")
    p_ss = p[idx + len("/screenshots/"):] if idx >= 0 else p
    # 1) results 런 폴더 기준
    cand = RESULTS_DIR / p
    if cand.exists():
        return cand
    # 2) screenshots 폴더 기준 (옛 형식 / expected 이미지)
    cand = SCREENSHOTS_DIR / p_ss
    if cand.exists():
        return cand
    return None


def _html_image_src(rel_path: str | None, html_dir: Path) -> str:
    """저장된 이미지 경로(RESULTS_DIR 또는 SCREENSHOTS_DIR 기준)를 HTML이 위치한
    디렉토리 기준 상대 경로로 변환. 파일이 존재하지 않으면 빈 문자열."""
    if not rel_path:
        return ""
    rp = str(rel_path).replace("\\", "/")
    # /screenshots/ 프리픽스가 포함된 경우 제거
    idx = rp.find("/screenshots/")
    if idx >= 0:
        rp_ss = rp[idx + len("/screenshots/"):]
    else:
        rp_ss = rp
    candidates = [
        RESULTS_DIR / rp,
        SCREENSHOTS_DIR / rp_ss,
    ]
    for abs_path in candidates:
        try:
            if abs_path.exists():
                try:
                    return abs_path.relative_to(html_dir).as_posix()
                except ValueError:
                    return os.path.relpath(str(abs_path), str(html_dir)).replace("\\", "/")
        except OSError:
            continue
    return ""


_HTML_STYLE = """
*,*::before,*::after { box-sizing: border-box; }
body { font-family: -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
  margin: 0; padding: 16px 20px; font-size: 13px; color: #1a1a2e; background: #f5f7fa; }
h1 { font-size: 20px; font-weight: 700; margin: 0 0 6px; color: #1a1a2e; letter-spacing: -0.3px; }
.meta { display: flex; flex-wrap: wrap; gap: 6px 16px; margin: 0 0 14px; color: #555; font-size: 12px; }
.meta .badge { display: inline-flex; align-items: center; gap: 4px; padding: 2px 10px;
  border-radius: 12px; font-weight: 600; font-size: 12px; }
.meta .badge.pass { background: #dcfce7; color: #15803d; }
.meta .badge.fail { background: #fee2e2; color: #b91c1c; }
.meta .badge.error { background: #ffedd5; color: #9a3412; }
.meta .stat { padding: 2px 8px; border-radius: 4px; background: #e8ecf1; font-weight: 500; }
.controls { position: sticky; top: 0; z-index: 20; background: #fff; padding: 10px 16px;
  border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); margin-bottom: 12px;
  display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }
.controls input[type="text"] { font-size: 12px; padding: 6px 10px; border: 1px solid #d1d5db;
  border-radius: 6px; min-width: 260px; outline: none; transition: border-color 0.15s; }
.controls input[type="text"]:focus { border-color: #3b82f6; box-shadow: 0 0 0 2px rgba(59,130,246,0.15); }
.controls button { font-size: 12px; padding: 6px 14px; border: none; border-radius: 6px;
  cursor: pointer; font-weight: 500; transition: all 0.15s; }
.controls .btn-primary { background: #3b82f6; color: #fff; }
.controls .btn-primary:hover { background: #2563eb; }
.controls .btn-secondary { background: #f3f4f6; color: #374151; border: 1px solid #d1d5db; }
.controls .btn-secondary:hover { background: #e5e7eb; }
.controls .count { margin-left: auto; color: #6b7280; font-size: 12px; }
/* Tabulator 테마 오버라이드 */
.tabulator { border: none; border-radius: 8px; overflow: hidden;
  box-shadow: 0 1px 4px rgba(0,0,0,0.08); font-size: 12px; background: #fff; }
.tabulator .tabulator-header { background: #1e3a5f; color: #fff; font-weight: 600;
  border-bottom: 2px solid #15294a; }
.tabulator .tabulator-header .tabulator-col { background: transparent; color: #fff;
  border-right: 1px solid rgba(255,255,255,0.1); }
.tabulator .tabulator-header .tabulator-col.tabulator-sortable:hover { background: rgba(255,255,255,0.1); }
.tabulator .tabulator-header .tabulator-col .tabulator-col-title { color: #fff; padding: 8px 6px;
  font-size: 11px; text-transform: uppercase; letter-spacing: 0.3px; }
.tabulator .tabulator-header .tabulator-col .tabulator-col-sorter { color: rgba(255,255,255,0.5); }
.tabulator .tabulator-header .tabulator-header-filter { padding: 4px 4px 6px; }
.tabulator .tabulator-header .tabulator-header-filter input,
.tabulator .tabulator-header .tabulator-header-filter select {
  font-size: 11px; padding: 4px 6px; border: 1px solid #d1d5db; border-radius: 4px;
  background: #fff; color: #374151; width: 100%; outline: none; }
.tabulator .tabulator-header .tabulator-header-filter input:focus,
.tabulator .tabulator-header .tabulator-header-filter select:focus {
  border-color: #60a5fa; box-shadow: 0 0 0 2px rgba(96,165,250,0.2); }
.tabulator .tabulator-tableholder { background: #fff; }
.tabulator-row { border-bottom: 1px solid #f0f0f0; }
.tabulator-row .tabulator-cell { padding: 4px 6px; border-right: 1px solid #f3f4f6; }
.tabulator-row.tabulator-row-even { background: #fafbfc; }
.tabulator-row:hover { background: #eff6ff !important; }
.tabulator-row.tabulator-row-even:hover { background: #eff6ff !important; }
/* Status 배지 */
.st-badge { display: inline-block; padding: 3px 10px; border-radius: 10px; font-size: 11px;
  font-weight: 700; letter-spacing: 0.3px; text-transform: uppercase; }
.st-badge.pass { background: #dcfce7; color: #15803d; }
.st-badge.fail { background: #fee2e2; color: #b91c1c; }
.st-badge.warning { background: #fef9c3; color: #854d0e; }
.st-badge.error { background: #ffedd5; color: #9a3412; }
.img-thumb { width: 160px; height: 118px; object-fit: contain; background: #fff;
  display: block; margin: 0 auto;
  cursor: pointer; border-radius: 4px; border: 1px solid #e5e7eb; transition: transform 0.15s; }
.img-thumb:hover { transform: scale(1.03); box-shadow: 0 2px 8px rgba(0,0,0,0.12); }
/* 로딩 오버레이 (테이블 빌드 전) */
.table-wrap { position: relative; }
.report-loading { position: absolute; inset: 0; min-height: 200px; background: rgba(255,255,255,0.92);
  display: flex; flex-direction: column; align-items: center; justify-content: center; gap: 14px;
  z-index: 30; }
.report-loading .spinner { width: 38px; height: 38px; border: 4px solid #dbeafe;
  border-top-color: #3b82f6; border-radius: 50%; animation: rk-spin 0.8s linear infinite; }
.report-loading-text { color: #475569; font-size: 14px; font-weight: 500; }
.report-loading-text b { color: #1e3a5f; }
@keyframes rk-spin { to { transform: rotate(360deg); } }
/* 이미지 프리뷰 */
.preview-overlay { position: fixed; inset: 0; background: rgba(0,0,0,0.88); display: none;
  align-items: center; justify-content: center; z-index: 9999; cursor: zoom-out; }
.preview-overlay.open { display: flex; }
.preview-overlay img { max-width: 95vw; max-height: 95vh; border-radius: 6px;
  box-shadow: 0 8px 40px rgba(0,0,0,0.4); }
@media print {
  body { margin: 8px; padding: 0; background: #fff; }
  .controls { display: none !important; }
  .report-loading { display: none !important; }
  .tabulator .tabulator-header .tabulator-header-filter { display: none !important; }
  .tabulator { box-shadow: none; border: 1px solid #ccc; height: auto !important;
    max-height: none !important; overflow: visible !important; }
  .tabulator .tabulator-tableholder { height: auto !important; max-height: none !important;
    overflow: visible !important; }
  .tabulator-row { page-break-inside: avoid; }
  .img-thumb { max-width: 140px; max-height: 110px; border: none; }
}
"""

# Tabulator 기반 리포트 초기화 스크립트 — 데이터는 window.__REPORT_DATA__ 전역에서 읽는다.
_HTML_SCRIPT = r"""
(function(){
  /* ---------- 셀 포맷터 ---------- */
  function statusFmt(cell){
    var v = (cell.getValue() || '').toString().toLowerCase();
    return '<span class="st-badge ' + v + '">' + v.toUpperCase() + '</span>';
  }
  function imgFmt(cell){
    var v = cell.getValue();
    if (!v) return '<span style="color:#bbb">—</span>';
    // width/height를 고정 예약 → lazy 이미지 로드 시 행 높이가 변하지 않아
    // 가상 스크롤이 튀지 않는다(예약 없으면 로드 전 0px→로드 후 커짐).
    return '<img class="img-thumb" width="160" height="118" loading="lazy" src="' + v + '" alt="">';
  }

  /* ---------- 고유값 수집 헬퍼 ---------- */
  function uniqueVals(data, field){
    var seen = {};
    var out = [];
    for (var i = 0; i < data.length; i++){
      var v = data[i][field];
      if (v == null || v === '') continue;
      var s = String(v);
      if (!seen[s]) { seen[s] = true; out.push(s); }
    }
    out.sort(function(a,b){ return a.localeCompare(b, undefined, {numeric:true}); });
    return out;
  }

  /* ---------- 드롭다운 헤더 필터 에디터 (리스트형) ---------- */
  function listEditor(field, placeholder){
    return function(cell, onRendered, success, cancel, editorParams){
      var vals = editorParams.values || [];
      var select = document.createElement('select');
      select.style.cssText = 'width:100%; font-size:11px; padding:3px 4px; border:1px solid #d1d5db; border-radius:4px; background:#fff; color:#374151; cursor:pointer;';
      var opt0 = document.createElement('option');
      opt0.value = ''; opt0.textContent = placeholder || '전체';
      select.appendChild(opt0);
      for (var i = 0; i < vals.length; i++){
        var opt = document.createElement('option');
        opt.value = vals[i]; opt.textContent = vals[i];
        select.appendChild(opt);
      }
      var cur = cell.getValue();
      if (cur) select.value = cur;
      select.addEventListener('change', function(){ success(select.value); });
      return select;
    };
  }

  /* ---------- 적응형 헤더 필터 ----------
     고유값이 많은 컬럼(timestamp/command 등)은 드롭다운 대신 텍스트 부분일치 입력으로
     전환한다. <select>에 수천~수만 옵션을 만들면 init/필터링이 멈추기 때문. */
  var MAX_DROPDOWN = 50;
  function filterFor(field, data, placeholder){
    var vals = uniqueVals(data, field);
    if (vals.length > MAX_DROPDOWN){
      return { headerFilter: "input", headerFilterFunc: "like",
               headerFilterPlaceholder: (placeholder || '검색') };
    }
    return { headerFilter: listEditor(field, placeholder),
             headerFilterParams: { values: vals }, headerFilterFunc: "=" };
  }

  /* ---------- 이미지 프리뷰 ---------- */
  function onImgClick(e){
    if (e.target && e.target.classList.contains('img-thumb')){
      document.getElementById('preview-img').src = e.target.src;
      document.getElementById('preview-overlay').classList.add('open');
    }
  }
  function closePreview(){
    document.getElementById('preview-overlay').classList.remove('open');
    document.getElementById('preview-img').src = '';
  }

  /* ---------- 컬럼 정의 ---------- */
  function buildColumns(data){
    var va = "middle";
    function col(base, field, placeholder){
      // base 컬럼 정의에 적응형 필터(드롭다운 또는 텍스트 입력)를 병합
      return Object.assign(base, filterFor(field, data, placeholder));
    }
    return [
      col({ title:"Time Stamp", field:"timestamp", width:150, vertAlign:va }, 'timestamp', '전체'),
      col({ title:"Cycle", field:"cycle", width:70, hozAlign:"center", vertAlign:va }, 'cycle', '전체'),
      col({ title:"Step", field:"step_id", width:70, hozAlign:"center", vertAlign:va }, 'step_id', '전체'),
      col({ title:"Device", field:"device", width:120, hozAlign:"center", vertAlign:va }, 'device', '전체'),
      col({ title:"Command", field:"command", widthGrow:2, vertAlign:va }, 'command', '전체'),
      { title:"Output", field:"output", widthGrow:3, vertAlign:va,
        formatter:function(cell){
          var v = cell.getValue();
          if (!v) return '<span style="color:#888">-</span>';
          var safe = String(v).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
          return '<pre style="margin:0;font-size:10px;font-family:monospace;white-space:pre-wrap;word-break:break-all;max-height:200px;overflow:auto">' + safe + '</pre>';
        },
        headerFilter:"input", headerFilterFunc:"like" },
      col({ title:"Remark", field:"description", widthGrow:2, vertAlign:va }, 'description', '전체'),
      { title:"Status", field:"status", width:90, hozAlign:"center", vertAlign:va, formatter:statusFmt,
        headerFilter:listEditor('status','전체'), headerFilterParams:{values:["pass","fail","warning","error"]}, headerFilterFunc:"=" },
      col({ title:"Delay", field:"delay", width:80, hozAlign:"center", vertAlign:va }, 'delay', '전체'),
      col({ title:"Duration", field:"duration", width:90, hozAlign:"center", vertAlign:va }, 'duration', '전체'),
      { title:"Expected", field:"expected_src", width:200, hozAlign:"center", vertAlign:va,
        formatter:imgFmt, headerSort:false },
      { title:"Actual", field:"actual_src", width:200, hozAlign:"center", vertAlign:va,
        formatter:imgFmt, headerSort:false },
    ];
  }

  /* ---------- 전역 검색 필터 ---------- */
  var TEXT_FIELDS = ["timestamp","cycle","step_id","device","command","output","description","status","delay","duration"];
  function globalFilter(row, params){
    var q = params.q;
    if (!q) return true;
    for (var i = 0; i < TEXT_FIELDS.length; i++){
      var v = row[TEXT_FIELDS[i]];
      if (v != null && String(v).toLowerCase().indexOf(q) !== -1) return true;
    }
    return false;
  }

  /* ---------- 로딩 오버레이 ---------- */
  function showLoading(n){
    var el = document.getElementById('report-loading');
    if (!el) return;
    var cnt = document.getElementById('report-loading-count');
    if (cnt) cnt.textContent = n ? (n.toLocaleString() + '건') : '';
    el.style.display = 'flex';
  }
  function hideLoading(){
    var el = document.getElementById('report-loading');
    if (el) el.style.display = 'none';
  }

  /* ---------- 압축 데이터 해제 ---------- */
  function decodeBase64ToBytes(b64){
    var bin = atob(b64 || '');
    var len = bin.length;
    var bytes = new Uint8Array(len);
    for (var i = 0; i < len; i++) bytes[i] = bin.charCodeAt(i);
    return bytes;
  }
  function loadReportData(){
    // gzip+base64로 임베드된 데이터를 브라우저에서 해제 → JSON 파싱.
    if (typeof DecompressionStream === 'undefined') {
      return Promise.reject(new Error('DecompressionStream 미지원 브라우저'));
    }
    var bytes = decodeBase64ToBytes(window.__RD__);
    var stream = new Blob([bytes]).stream().pipeThrough(new DecompressionStream('gzip'));
    return new Response(stream).text().then(function(txt){ return JSON.parse(txt); });
  }

  /* ---------- 초기화 ---------- */
  function init(){
    showLoading(0);  // 해제 전부터 로딩 표시
    loadReportData().then(function(payload){
      window.__REPORT_DATA__ = payload;
      var data = (payload && payload.rows) || [];
      document.getElementById('filter-total').textContent = data.length;
      // 한 틱 뒤 테이블 빌드(로딩 오버레이를 먼저 페인트하도록)
      showLoading(data.length);
      setTimeout(function(){ buildTable(data); }, 0);
    }).catch(function(err){
      hideLoading();
      var el = document.getElementById('results-table');
      if (el) el.innerHTML = '<div style="padding:48px;text-align:center;color:#b91c1c;font-size:14px">'
        + '리포트 데이터를 여는 데 실패했습니다. 최신 Chrome/Edge/Firefox에서 열어 주세요.'
        + '<div style="color:#888;font-size:12px;margin-top:8px">(' + (err && err.message ? err.message : err) + ')</div></div>';
    });
  }

  function buildTable(data){
    var opts = {
      data: data,
      columns: buildColumns(data),
      // fitColumns: 데이터 기반 폭 측정을 생략해 대용량에서 빌드가 더 빠르다.
      layout: "fitColumns",
      // 대용량(수천~수만 행) 리포트 대응: 가상 스크롤로 보이는 행만 렌더.
      // 'basic'은 전 행을 DOM에 마운트해 13744행 같은 경우 브라우저가 멈춘다.
      renderVertical: "virtual",
      placeholder: "표시할 결과가 없습니다",
      headerSortClickElement: "icon",
      rowHeight: false,
      cellVertAlign: "middle",
      // PDF 저장은 table.print('all')로 전 행을 별도 렌더 → 가상 스크롤과 무관하게 전체 출력
      printAsHtml: true,
      printRowRange: "all",
      printConfig: { columnHeaders: true },
    };
    // 대용량은 고정 height(가상 스크롤이 가장 안정적), 소량은 maxHeight로 내용 맞춤.
    if (data.length > 200) opts.height = "calc(100vh - 175px)";
    else opts.maxHeight = "calc(100vh - 175px)";
    var table = new Tabulator("#results-table", opts);
    window.__table = table;
    // 빌드가 끝나면 로딩 표시 제거
    table.on("tableBuilt", hideLoading);
    table.on("renderComplete", hideLoading);

    table.on("dataFiltered", function(filters, rows){
      document.getElementById('filter-visible').textContent = rows.length;
    });

    // 전역 검색
    var gEl = document.getElementById('filter-text');
    gEl.addEventListener('input', function(){
      var v = (gEl.value || '').trim().toLowerCase();
      table.removeFilter(globalFilter);
      if (v) table.addFilter(globalFilter, { q: v });
    });

    // 필터 초기화
    document.getElementById('filter-reset').addEventListener('click', function(){
      table.clearHeaderFilter();
      gEl.value = '';
      table.removeFilter(globalFilter);
    });

    // PDF 저장 — Tabulator 내장 print로 전 행 출력(가상 스크롤이어도 전체가 나옴)
    document.getElementById('pdf-btn').addEventListener('click', function(){ table.print("all", true); });

    // 이미지 프리뷰
    document.getElementById('results-table').addEventListener('click', onImgClick);
    document.getElementById('preview-overlay').addEventListener('click', closePreview);
    document.addEventListener('keydown', function(e){ if (e.key === 'Escape') closePreview(); });
  }

  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', init);
  else init();
})();
"""


def _build_html_report(data: dict, output_path: Path, steps_iter=None) -> str:
    """Tabulator 기반 경량 HTML 리포트.

    데이터는 gzip+base64로 압축해 window.__RD__에 임베드되고, 브라우저에서
    DecompressionStream으로 해제 후 window.__REPORT_DATA__로 사용한다.
    Tabulator가 열별 필터/정렬/검색/이미지 썸네일을 모두 렌더링한다.
    라이브러리 파일은 /static/tabulator/ 에서 서빙 (별도 복사 불필요).
    export-bundle ZIP에만 assets/로 포함된다.

    steps_iter: 스텝 dict의 이터러블(예: NDJSON 라인 제너레이터). 주어지면
    data["step_results"] 대신 이걸 흘려 읽는다. 행 전량을 메모리에 올리지 않고
    gzip 스트림에 직접 압축하므로, 수만 행(장시간 aging test)에서도 피크 메모리가
    압축 바이트 크기로 고정된다.
    """
    html_dir = output_path.parent

    def e(v) -> str:
        return _html.escape("" if v is None else str(v))

    scenario_name = data.get("scenario_name", "")
    status = data.get("status", "")
    total_steps = data.get("total_steps", 0)
    total_repeat = data.get("total_repeat", 1)
    passed = data.get("passed_steps", 0)
    failed = data.get("failed_steps", 0)
    warned = data.get("warning_steps", 0)
    errored = data.get("error_steps", 0)
    started_at = data.get("started_at", "")
    finished_at = data.get("finished_at", "")

    def _fmt_ts(iso: str) -> str:
        try:
            from datetime import datetime as _dt
            ts = _dt.fromisoformat(iso.replace("Z", "+00:00"))
            return ts.astimezone().strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return iso or ""

    # Tabulator에 넘길 행 dict 1개 구성 (스텝 dict → 표시용 행)
    def _row_for(sr: dict) -> dict:
        duration_ms = sr.get("execution_time_ms", 0) or 0
        delay_ms = sr.get("delay_ms", 0) or 0
        dur_str = f"{duration_ms}ms" if duration_ms < 1000 else f"{duration_ms / 1000:.1f}s"
        delay_str = f"{delay_ms}ms" if delay_ms else "-"
        exp_src = _html_image_src(sr.get("expected_image"), html_dir)
        act_src = _html_image_src(
            sr.get("actual_annotated_image") or sr.get("actual_image"), html_dir
        )
        return {
            "timestamp": _fmt_ts(sr.get("timestamp", started_at)),
            "cycle": sr.get("repeat_index", 1),
            "step_id": sr.get("step_id", ""),
            "device": sr.get("device_id", ""),
            "command": sr.get("command", ""),
            # 모듈 명령(cmd, adb_send 등)의 실행 출력값. 빈 값은 명령 표시와 별개 컬럼에서 - 처리.
            "output": sr.get("message", ""),
            "description": sr.get("description", ""),
            "status": sr.get("status", ""),
            "delay": delay_str,
            "duration": dur_str,
            "expected_src": exp_src or "",
            "actual_src": act_src or "",
        }

    # 데이터는 gzip→base64로 압축 임베드 → 대용량 리포트의 파일 크기를 10~20배 줄인다.
    # (JSON은 반복이 많아 압축률이 높음.) base64 알파벳엔 < " \ 가 없어
    # <script> 안 JS 문자열에 그대로 안전(별도 이스케이프 불필요).
    # 행을 리스트로 모으거나 비압축 JSON 문자열을 통째로 만들지 않고 gzip 스트림에
    # 직접 흘려 쓴다 → 수만 행에서도 피크 메모리가 압축 바이트 크기로 고정.
    steps_source = steps_iter if steps_iter is not None else data.get("step_results", [])
    _buf = io.BytesIO()
    with gzip.GzipFile(fileobj=_buf, mode="wb", compresslevel=9, mtime=0) as _gz:
        _gz.write(b'{"scenario_name":')
        _gz.write(json.dumps(scenario_name, ensure_ascii=False).encode("utf-8"))
        _gz.write(b',"status":')
        _gz.write(json.dumps(status, ensure_ascii=False).encode("utf-8"))
        _gz.write(b',"total_repeat":')
        _gz.write(json.dumps(total_repeat).encode("utf-8"))
        _gz.write(b',"rows":[')
        _first = True
        for sr in steps_source:
            if not _first:
                _gz.write(b",")
            _gz.write(json.dumps(_row_for(sr), ensure_ascii=False).encode("utf-8"))
            _first = False
        _gz.write(b"]}")
    payload_b64 = base64.b64encode(_buf.getvalue()).decode("ascii")

    parts: list[str] = []
    parts.append("<!DOCTYPE html>")
    parts.append('<html lang="ko"><head><meta charset="utf-8">')
    parts.append(f"<title>{e(scenario_name)} - Test Report</title>")
    parts.append('<script>var _tBase = location.protocol==="file:" ? "../../app/static/tabulator/" : "/static/tabulator/";</script>')
    parts.append('<script>document.write(\'<link rel="stylesheet" href="\'+_tBase+\'tabulator_simple.min.css">\');</script>')
    parts.append(f"<style>{_HTML_STYLE}</style>")
    parts.append("</head><body>")
    parts.append(f"<h1>{e(scenario_name)}</h1>")
    status_cls = status if status in ("pass", "fail", "error") else ""
    parts.append('<div class="meta">')
    parts.append(f'<span class="badge {status_cls}">{e(status.upper())}</span>')
    parts.append(f'<span class="stat">Step: {total_steps}</span>')
    parts.append(f'<span class="stat">Repeat: {total_repeat}</span>')
    parts.append(f'<span class="stat" style="background:#dcfce7;color:#15803d">Pass: {passed}</span>')
    if failed:
        parts.append(f'<span class="stat" style="background:#fee2e2;color:#b91c1c">Fail: {failed}</span>')
    if warned:
        parts.append(f'<span class="stat" style="background:#fef9c3;color:#854d0e">Warn: {warned}</span>')
    if errored:
        parts.append(f'<span class="stat" style="background:#ffedd5;color:#9a3412">Error: {errored}</span>')
    parts.append(f"<span>{e(_fmt_ts(started_at))} ~ {e(_fmt_ts(finished_at))}</span>")
    parts.append("</div>")

    # Frame_Check 측정 결과 — 녹화 영상 프레임 분석 (시작점 → 타겟 이미지 경과 시간)
    fc_results = data.get("frame_check_results") or []
    if fc_results:
        parts.append('<h2 style="font-size:14px;margin:12px 0 6px">Frame Check — 동작 시간 측정</h2>')
        parts.append('<table style="border-collapse:collapse;font-size:12px;margin-bottom:12px">')
        parts.append(
            "<tr>" + "".join(
                f'<th style="border:1px solid #d1d5db;padding:3px 8px;background:#f3f4f6">{h}</th>'
                for h in ("Cycle", "Pair", "시작 기준", "시작 시각", "타겟 시각",
                          "경과 시간(ms)", "Score(시작/타겟)", "매치", "Status", "Clip", "Message")
            ) + "</tr>"
        )
        for fc in fc_results:
            mode = fc.get("start_mode")
            mode_str = "이미지 등장" if mode == "image" else ("스텝 실행" if mode == "function" else "-")
            elapsed = fc.get("elapsed_ms")
            elapsed_str = f"<b>{elapsed:,.0f}</b>" if isinstance(elapsed, (int, float)) else "-"
            sc = fc.get("start_score")
            tc = fc.get("target_score")
            score_str = f"{sc if sc is not None else '-'} / {tc if tc is not None else '-'}"
            st = str(fc.get("status", ""))
            st_color = "#15803d" if st == "ok" else "#b91c1c"
            # 시작/타겟 시각 — 녹화 started_at + 영상 오프셋의 wall-clock (없으면 영상 ms 폴백)
            def _t(fc_row, time_key, ms_key):
                v = fc_row.get(time_key)
                if v:
                    return e(v)
                ms = fc_row.get(ms_key)
                return f"{e(ms)} ms" if ms is not None else "-"
            # 클립/매치 이미지는 result.html 과 같은 폴더의 recordings/ 에 있어 상대 링크로 동작
            # (서버 서빙 /results-files/... 과 file:// 더블클릭 양쪽 모두).
            clip_rel = str(fc.get("clip") or "")
            clip_link = (
                f'<a href="{e(clip_rel)}" target="_blank">&#9654; 재생</a>' if clip_rel else "-"
            )
            match_rel = str(fc.get("match_image") or "")
            match_link = (
                f'<a href="{e(match_rel)}" target="_blank">'
                f'<img src="{e(match_rel)}" style="height:28px;border-radius:3px;vertical-align:middle" '
                f'alt="match"></a>' if match_rel else "-"
            )
            cells = [
                f"R{e(fc.get('iteration', ''))}",
                e(fc.get("pair_index", "-")),
                mode_str,
                _t(fc, "start_time", "start_video_ms"),
                _t(fc, "target_time", "target_video_ms"),
                elapsed_str,
                e(score_str),
                match_link,
                f'<span style="color:{st_color};font-weight:600">{e(st.upper())}</span>',
                clip_link,
                e(fc.get("message", "")),
            ]
            parts.append(
                "<tr>" + "".join(
                    f'<td style="border:1px solid #d1d5db;padding:3px 8px">{c}</td>' for c in cells
                ) + "</tr>"
            )
        parts.append("</table>")

    # 상단 컨트롤 — 전역 검색, 필터 초기화, PDF 저장
    parts.append('<div class="controls">')
    parts.append('<input id="filter-text" type="text" placeholder="전체 검색 (모든 열)">')
    parts.append('<button id="filter-reset" class="btn-secondary" type="button">필터 초기화</button>')
    parts.append('<button id="pdf-btn" class="btn-primary" type="button">PDF 저장</button>')
    parts.append('<span class="count"><b id="filter-visible">0</b> / <b id="filter-total">0</b></span>')
    parts.append("</div>")

    # Tabulator 렌더 타겟 + 로딩 오버레이 (빌드 완료 전까지 표시)
    parts.append('<div class="table-wrap">')
    parts.append('<div id="results-table"></div>')
    parts.append(
        '<div id="report-loading" class="report-loading">'
        '<div class="spinner"></div>'
        '<div class="report-loading-text">데이터 로딩 중… '
        '<b id="report-loading-count"></b></div>'
        '</div>'
    )
    parts.append("</div>")

    # 이미지 프리뷰 오버레이
    parts.append('<div id="preview-overlay" class="preview-overlay"><img id="preview-img" src="" alt=""></div>')

    # 데이터 임베드(gzip+base64) + 라이브러리 로드 + 초기화
    parts.append(f'<script>window.__RD__="{payload_b64}";</script>')
    parts.append('<script>document.write(\'<script src="\'+_tBase+\'tabulator.min.js"><\\/script>\');</script>')
    parts.append(f"<script>{_HTML_SCRIPT}</script>")
    parts.append("</body></html>")
    return "".join(parts)


# Excel 썸네일 한계 — 원본 전체 해상도를 임베드하면 파일이 수 GB로 폭증한다.
# (XlImage.width/height는 '표시' 크기만 바꿀 뿐 저장 바이트는 원본 그대로이기 때문.)
_EXCEL_THUMB_MAX = (300, 230)              # 썸네일 픽셀 상한
_EXCEL_IMG_BUDGET = 150 * 1024 * 1024      # 누적 임베드 이미지 바이트 상한(초과 시 경로 텍스트)


def _excel_thumb(path):
    """이미지를 축소한 JPEG 썸네일 XlImage와 바이트 수를 반환. 실패 시 (None, 0).

    원본 전체 바이트 대신 축소본만 임베드해 xlsx 용량/생성시간 폭증을 막는다.
    """
    try:
        from openpyxl.drawing.image import Image as XlImage
        from PIL import Image as _PILImage
    except Exception:
        return None, 0
    try:
        with _PILImage.open(str(path)) as im:
            im = im.convert("RGB")
            im.thumbnail(_EXCEL_THUMB_MAX)
            buf = io.BytesIO()
            im.save(buf, format="JPEG", quality=70)
        nbytes = buf.tell()
        buf.seek(0)
        return XlImage(buf), nbytes
    except Exception:
        return None, 0


def _build_excel_workbook(data: dict, filepath: Path = None, progress=None):
    """Build an openpyxl Workbook from result data. Reusable by settings router.

    progress(done:int, total:int): 스텝 단위 진행 콜백(선택). 이미지 임베드가
    무거우므로 내보내기 진행률 표시에 사용한다.

    이미지는 원본이 아니라 축소 썸네일을 임베드하고, 누적 바이트가 상한을 넘으면
    경로 텍스트로 대체한다(대용량 결과의 xlsx 용량 폭증 방지).
    """
    import openpyxl
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _xl_text(v):
        # openpyxl은 제어문자(예: 시리얼/SSH 출력의 ANSI escape \x1b)가 셀 값에
        # 있으면 IllegalCharacterError를 던진다 — 제거 후 기록.
        if v is None:
            return ""
        if not isinstance(v, str):
            v = str(v)
        return ILLEGAL_CHARACTERS_RE.sub("", v)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Report"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    desc_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    desc_font = Font(color="44546A", size=9)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    vcenter = Alignment(vertical="center")
    vcenter_wrap = Alignment(vertical="center", wrap_text=True)

    col_headers = [
        "Time Stamp", "TOTAL TC REPEAT", "CURRENT TC REPEAT",
        "STEP INDEX", "Device", "Command", "Output", "Remark", "Status", "DELAY", "DURATION",
        "Expected Image", "Actual Image",
    ]
    col_descs = [
        "실행된 날짜/시간", "총 repeat", "현재 cycle",
        "스탭 순서", "장치", "action", "명령 실행 출력값", "설명", "pass, fail, error, jump", "설정한 딜레이", "실제 걸린 시간",
        "기대 이미지", "비교 이미지 (annotated)",
    ]
    col_widths = [22, 16, 18, 12, 16, 30, 40, 30, 12, 12, 14, 30, 30]

    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for ci, d in enumerate(col_descs, start=1):
        cell = ws.cell(row=2, column=ci, value=d)
        cell.font = desc_font
        cell.fill = desc_fill
        cell.alignment = center
        cell.border = thin_border

    total_repeat = data.get("total_repeat", 1)
    img_row_height = 120

    _steps = data.get("step_results", [])
    _total_steps = len(_steps)
    _img_bytes = 0  # 누적 임베드 이미지 바이트(상한 초과 시 이미지 대신 경로 텍스트)
    for ri, sr in enumerate(_steps, start=3):
        if progress is not None:
            try:
                progress(ri - 2, _total_steps)
            except Exception:
                pass
        # 값이 null 로 저장된 스텝(중단/에러)도 있으므로 None 방어 필수
        status = sr.get("status") or ""
        timestamp = sr.get("timestamp") or data.get("started_at") or ""
        command = _xl_text(sr.get("command"))
        output = _xl_text(sr.get("message"))
        delay_ms = sr.get("delay_ms") or 0
        duration_ms = sr.get("execution_time_ms") or 0

        try:
            from datetime import datetime as _dt, timezone as _tz
            ts = _dt.fromisoformat(timestamp.replace("Z", "+00:00"))
            ts_local = ts.astimezone()  # 시스템 로컬 시간대로 변환
            ts_str = ts_local.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            ts_str = timestamp or ""

        dur_str = f"{duration_ms}ms" if duration_ms < 1000 else f"{duration_ms / 1000:.1f}s"
        delay_str = f"{delay_ms}ms" if delay_ms and delay_ms >= 1000 else (f"{delay_ms}ms" if delay_ms else "-")

        ws.cell(row=ri, column=1, value=ts_str).border = thin_border
        ws.cell(row=ri, column=1).alignment = center
        ws.cell(row=ri, column=2, value=total_repeat).border = thin_border
        ws.cell(row=ri, column=2).alignment = center
        ws.cell(row=ri, column=3, value=sr.get("repeat_index", 1)).border = thin_border
        ws.cell(row=ri, column=3).alignment = center
        ws.cell(row=ri, column=4, value=sr.get("step_id", "")).border = thin_border
        ws.cell(row=ri, column=4).alignment = center
        ws.cell(row=ri, column=5, value=_xl_text(sr.get("device_id"))).border = thin_border
        ws.cell(row=ri, column=5).alignment = center
        ws.cell(row=ri, column=6, value=command).border = thin_border
        ws.cell(row=ri, column=6).alignment = vcenter_wrap
        # Output (모듈 명령 실행 출력값) — 모듈 명령이 아닌 일반 스텝은 빈 값.
        ws.cell(row=ri, column=7, value=output).border = thin_border
        ws.cell(row=ri, column=7).alignment = vcenter_wrap
        ws.cell(row=ri, column=8, value=_xl_text(sr.get("description"))).border = thin_border
        ws.cell(row=ri, column=8).alignment = vcenter_wrap
        status_cell = ws.cell(row=ri, column=9, value=status.upper())
        status_cell.border = thin_border
        status_cell.alignment = center
        if status == "pass":
            status_cell.fill = pass_fill
        elif status == "fail":
            status_cell.fill = fail_fill
        elif status == "warning":
            status_cell.fill = warn_fill
        elif status == "error":
            status_cell.fill = error_fill
        ws.cell(row=ri, column=10, value=delay_str).border = thin_border
        ws.cell(row=ri, column=10).alignment = center
        ws.cell(row=ri, column=11, value=dur_str).border = thin_border
        ws.cell(row=ri, column=11).alignment = center

        exp_path = _resolve_image_path(sr.get("expected_image"))
        ws.cell(row=ri, column=12).border = thin_border
        ws.cell(row=ri, column=12).alignment = center
        if exp_path:
            thumb, nb = _excel_thumb(exp_path) if _img_bytes < _EXCEL_IMG_BUDGET else (None, 0)
            if thumb is not None:
                try:
                    thumb.width = 180
                    thumb.height = 140
                    ws.add_image(thumb, f"L{ri}")
                    ws.row_dimensions[ri].height = img_row_height
                    _img_bytes += nb
                except Exception:
                    ws.cell(row=ri, column=12, value=str(sr.get("expected_image", "")))
            else:
                ws.cell(row=ri, column=12, value=str(sr.get("expected_image", "")))

        act_img_path = sr.get("actual_annotated_image") or sr.get("actual_image")
        act_path = _resolve_image_path(act_img_path)
        ws.cell(row=ri, column=13).border = thin_border
        ws.cell(row=ri, column=13).alignment = center
        if act_path:
            thumb, nb = _excel_thumb(act_path) if _img_bytes < _EXCEL_IMG_BUDGET else (None, 0)
            if thumb is not None:
                try:
                    thumb.width = 180
                    thumb.height = 140
                    ws.add_image(thumb, f"M{ri}")
                    if ws.row_dimensions[ri].height is None or ws.row_dimensions[ri].height < img_row_height:
                        ws.row_dimensions[ri].height = img_row_height
                    _img_bytes += nb
                except Exception:
                    ws.cell(row=ri, column=13, value=str(act_img_path or ""))
            else:
                ws.cell(row=ri, column=13, value=str(act_img_path or ""))

    return wb


@router.get("/export/{filename:path}")
async def export_result_excel(filename: str):
    """Export a test result as Excel (.xlsx) — download to browser."""
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")

    def _build() -> io.BytesIO:
        data = json.loads(filepath.read_text(encoding="utf-8"))
        wb = _build_excel_workbook(data, filepath)
        b = io.BytesIO()
        wb.save(b)
        b.seek(0)
        # 런 폴더면 result.xlsx 를 폴더에도 남긴다. 재생 저장 시점에는 이미지 임베드가
        # 무거워 엑셀을 만들지 않으므로, 결과 폴더에 result.xlsx 가 없다는 문의가 반복됐다.
        # (같은 바이트를 재사용 — wb.save 를 두 번 하지 않는다)
        if filepath.name == "result.json" and filepath.parent != RESULTS_DIR:
            try:
                (filepath.parent / "result.xlsx").write_bytes(b.getvalue())
            except Exception as e:
                logger.warning("result.xlsx 폴더 저장 실패: %s", e)
        return b

    try:
        # 대용량 결과의 JSON 파싱 + 엑셀 생성은 스레드로 — 이벤트 루프 보호
        buf = await asyncio.to_thread(_build)
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    except Exception as e:
        # 원인 불명 500 방지 — 프론트 에러 메시지와 backend.log 양쪽에 실제 원인을 남긴다
        logger.exception("Excel export failed: %s", filename)
        raise HTTPException(status_code=500, detail=f"Excel 생성 실패: {type(e).__name__}: {e}")

    export_name = Path(filename.replace(".json", ".xlsx")).name
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(export_name)},
    )


def _export_bundle_sync(filename: str, export_path: str, progress=None) -> dict:
    """결과 내보내기의 동기(블로킹) 본문 — 스레드에서 실행해 이벤트 루프를 막지 않는다.

    리포트 재생성(이미지 임베드 엑셀 포함)·ZIP 압축이 무거우므로 분리.
    progress(percent:int, phase:str): 전체 진행률 콜백(선택).

    반환:
      - export_path 지정: {"mode": "saved", "path", "folder", "size"}
      - 다운로드:        {"mode": "download", "zip_path", "folder", "size"}
    """
    def _p(pct: int, phase: str) -> None:
        if progress is not None:
            try:
                progress(pct, phase)
            except Exception:
                pass

    # 엑셀(이미지 임베드)은 5~78%, ZIP 압축은 82~99% 구간에 매핑한다.
    def _excel_prog(done, total):
        if total:
            _p(5 + int(73 * done / total), "Excel 생성 중")

    def _zip_prog(done, total):
        if total:
            _p(82 + int(17 * done / total), "압축 중")

    _p(1, "준비 중")
    filepath = RESULTS_DIR / filename

    # 런 폴더인지 레거시인지 판별
    if filepath.name == "result.json" and filepath.parent != RESULTS_DIR:
        run_dir = filepath.parent
        folder_name = run_dir.name
        # 내보내기 시점에 result.html / result.xlsx 를 항상 최신 코드로 재생성한다.
        # (저장 당시 옛 버전으로 구워진 리포트도 최신 렌더링/포맷으로 갱신됨)
        try:
            data = json.loads(filepath.read_text(encoding="utf-8"))
        except Exception:
            data = None
        if data is not None:
            try:
                html_path = run_dir / "result.html"
                html_path.write_text(_build_html_report(data, html_path), encoding="utf-8")
                _p(3, "HTML 생성 완료")
            except Exception as e:
                logger.warning("HTML report regeneration failed: %s", e)
            try:
                wb = _build_excel_workbook(data, filepath, progress=_excel_prog)
                wb.save(str(run_dir / "result.xlsx"))
            except Exception as e:
                logger.warning("Excel report regeneration failed: %s", e)
        _p(80, "압축 준비 중")
    else:
        # 레거시: 임시 폴더에 결과물 수집
        data = json.loads(filepath.read_text(encoding="utf-8"))
        scenario_name = data.get("scenario_name", "unknown")
        started_at = data.get("started_at", "")
        try:
            dt = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
            ts = dt.strftime("%Y%m%d_%H%M%S")
        except Exception:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = scenario_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
        folder_name = f"{ts}_{safe_name}"

        run_dir = Path(tempfile.mkdtemp()) / folder_name
        run_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(filepath), str(run_dir / filepath.name))

        # Excel 생성
        try:
            wb = _build_excel_workbook(data, filepath, progress=_excel_prog)
            wb.save(str(run_dir / filepath.name.replace(".json", ".xlsx")))
        except Exception as e:
            logger.warning("Excel report generation failed: %s", e)

        # HTML 리포트 생성 (레거시도 내보내기 시점에 최신 코드로 생성)
        try:
            html_path = run_dir / "result.html"
            html_path.write_text(_build_html_report(data, html_path), encoding="utf-8")
        except Exception as e:
            logger.warning("HTML report generation failed: %s", e)

        # 웹캠 녹화 복사 (webm + mp4)
        _p(80, "녹화 수집 중")
        base = filename.replace(".json", "")
        if RECORDINGS_DIR.is_dir():
            for pattern in (f"{base}_webcam_*.webm", f"{base}_webcam_*.mp4"):
                for rec in sorted(RECORDINGS_DIR.glob(pattern)):
                    try:
                        shutil.copy2(str(rec), str(run_dir / rec.name))
                    except Exception:
                        pass

    # result.html이 /static/tabulator/ 절대경로를 참조하므로, ZIP 배포용으로
    # assets/를 런 폴더에 임시 복사 + HTML 내 경로를 상대경로로 패치한다.
    _tabulator_src = Path(__file__).resolve().parent.parent / "static" / "tabulator"
    _tmp_assets_dir = run_dir / "assets"
    _patched_html = False
    if _tabulator_src.is_dir():
        _tmp_assets_dir.mkdir(exist_ok=True)
        for _tf in ("tabulator.min.js", "tabulator_simple.min.css"):
            _src = _tabulator_src / _tf
            _dst = _tmp_assets_dir / _tf
            if _src.is_file() and not _dst.exists():
                shutil.copy2(str(_src), str(_dst))
        _html_file = run_dir / "result.html"
        if _html_file.is_file():
            _htxt = _html_file.read_text(encoding="utf-8")
            # 프로토콜 감지 로직을 ./assets/ 고정으로 교체
            _htxt_new = _htxt.replace(
                'var _tBase = location.protocol==="file:" ? "../../app/static/tabulator/" : "/static/tabulator/";',
                'var _tBase = "./assets/";',
            )
            if _htxt_new != _htxt:
                _html_file.write_text(_htxt_new, encoding="utf-8")
                _patched_html = True

    # ZIP 압축
    try:
        if export_path:
            zip_path = Path(export_path)
            if zip_path.is_dir():
                zip_path = zip_path / f"{folder_name}.zip"
            zip_path.parent.mkdir(parents=True, exist_ok=True)
            _zip_directory(run_dir, zip_path, progress=_zip_prog)
            return {"mode": "saved", "path": str(zip_path), "folder": folder_name,
                    "size": zip_path.stat().st_size}
        else:
            # 다운로드: 임시 ZIP 파일로 저장 → 다운로드 엔드포인트가 서빙 후 삭제
            fd, tmp_zip = tempfile.mkstemp(suffix=".zip", prefix="rk_export_")
            os.close(fd)
            _zip_directory(run_dir, Path(tmp_zip), progress=_zip_prog)
            return {"mode": "download", "zip_path": tmp_zip, "folder": folder_name,
                    "size": Path(tmp_zip).stat().st_size}
    finally:
        # ZIP용 임시 assets 정리 + HTML 경로 복원 (런 폴더가 원본이면 패치 원복)
        if _patched_html:
            _html_file = run_dir / "result.html"
            if _html_file.is_file():
                _htxt = _html_file.read_text(encoding="utf-8")
                _html_file.write_text(
                    _htxt.replace(
                        'var _tBase = "./assets/";',
                        'var _tBase = location.protocol==="file:" ? "../../app/static/tabulator/" : "/static/tabulator/";',
                    ),
                    encoding="utf-8",
                )
        if _tmp_assets_dir.is_dir() and _tabulator_src.is_dir():
            shutil.rmtree(str(_tmp_assets_dir), ignore_errors=True)


# ---------- 내보내기 백그라운드 잡 (진행률 폴링) ----------
_EXPORT_JOBS: dict[str, dict] = {}
_EXPORT_JOBS_LOCK = threading.Lock()
_EXPORT_JOB_TTL = 3600  # 완료/실패 잡 보관 시간(초)


def _cleanup_export_jobs() -> None:
    """오래된 완료/실패 잡과 임시 ZIP 파일 정리."""
    now = time.monotonic()
    with _EXPORT_JOBS_LOCK:
        stale = [
            jid for jid, j in _EXPORT_JOBS.items()
            if j.get("status") in ("done", "error")
            and now - j.get("finished", now) > _EXPORT_JOB_TTL
        ]
        for jid in stale:
            j = _EXPORT_JOBS.pop(jid, None)
            if j and j.get("zip_path"):
                try:
                    os.unlink(j["zip_path"])
                except OSError:
                    pass


def _run_export_job(job_id: str, filename: str, export_path: str) -> None:
    """백그라운드 스레드에서 번들을 생성하며 잡 진행률을 갱신한다."""
    def _prog(pct, phase):
        with _EXPORT_JOBS_LOCK:
            j = _EXPORT_JOBS.get(job_id)
            if j:
                j["percent"] = max(j.get("percent", 0), min(99, int(pct)))
                j["phase"] = phase

    try:
        result = _export_bundle_sync(filename, export_path, _prog)
        with _EXPORT_JOBS_LOCK:
            j = _EXPORT_JOBS.get(job_id)
            if j:
                j.update(result)
                j["percent"] = 100
                j["phase"] = "완료"
                j["status"] = "done"
                j["finished"] = time.monotonic()
    except Exception as e:
        logger.exception("export bundle job failed: %s", filename)
        with _EXPORT_JOBS_LOCK:
            j = _EXPORT_JOBS.get(job_id)
            if j:
                j["status"] = "error"
                j["error"] = str(e)
                j["finished"] = time.monotonic()


@router.post("/export-bundle/{filename:path}")
async def export_result_bundle(filename: str, export_path: str = ""):
    """결과 내보내기 시작 — 백그라운드 잡으로 result.html/result.xlsx 재생성 + ZIP을
    처리하고 job_id를 즉시 반환한다.

    진행률: GET /api/results/export-job/{job_id}
    다운로드: GET /api/results/export-job/{job_id}/download (브라우저 다운로드 모드)

    Args:
        export_path: 저장 경로. 빈 값이면 브라우저 다운로드.
    """
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")

    _cleanup_export_jobs()
    job_id = uuid.uuid4().hex
    with _EXPORT_JOBS_LOCK:
        _EXPORT_JOBS[job_id] = {
            "status": "running", "percent": 0, "phase": "준비 중",
            "error": None, "mode": None, "folder": "", "created": time.monotonic(),
        }
    threading.Thread(
        target=_run_export_job, args=(job_id, filename, export_path), daemon=True
    ).start()
    return {"job_id": job_id}


@router.get("/export-job/{job_id}")
async def export_job_status(job_id: str):
    """내보내기 잡 진행률 조회."""
    with _EXPORT_JOBS_LOCK:
        j = _EXPORT_JOBS.get(job_id)
        if not j:
            raise HTTPException(status_code=404, detail="Job not found")
        return {
            "status": j["status"],
            "percent": j.get("percent", 0),
            "phase": j.get("phase", ""),
            "error": j.get("error"),
            "mode": j.get("mode"),
            "folder": j.get("folder", ""),
            "size": j.get("size"),
            "path": j.get("path"),  # saved 모드 저장 경로
        }


@router.get("/export-job/{job_id}/download")
async def export_job_download(job_id: str):
    """완료된 내보내기 잡의 ZIP 다운로드(브라우저 다운로드 모드). 전송 후 임시파일·잡 정리."""
    with _EXPORT_JOBS_LOCK:
        j = _EXPORT_JOBS.get(job_id)
    if not j or j.get("status") != "done" or j.get("mode") != "download":
        raise HTTPException(status_code=404, detail="Export not ready")
    zip_path = j.get("zip_path")
    if not zip_path or not Path(zip_path).exists():
        raise HTTPException(status_code=404, detail="Export file missing")
    folder = j.get("folder", "export")

    def _after():
        with _EXPORT_JOBS_LOCK:
            jj = _EXPORT_JOBS.pop(job_id, None)
        try:
            if jj and jj.get("zip_path"):
                os.unlink(jj["zip_path"])
        except OSError:
            pass

    return FileResponse(
        zip_path,
        media_type="application/zip",
        headers={"Content-Disposition": _content_disposition(f"{folder}.zip")},
        background=BackgroundTask(_after),
    )


@router.post("/regenerate-html/{filename:path}")
def regenerate_result_html(filename: str):
    """결과의 result.html을 현재 코드로 재생성한다(상세 모달 'HTML 생성' 버튼).

    저장 당시 옛 버전으로 구워진 리포트를 최신 렌더링/압축/포맷으로 갱신할 때 사용.
    sync def라 FastAPI가 스레드풀에서 실행 → 이벤트 루프를 막지 않는다.
    """
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")
    try:
        data = json.loads(filepath.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"결과 JSON 읽기 실패: {e}")

    html_path = _result_html_path(filepath)
    try:
        html_path.write_text(_build_html_report(data, html_path), encoding="utf-8")
    except Exception as e:
        logger.exception("HTML report regeneration failed: %s", filename)
        raise HTTPException(status_code=500, detail=f"HTML 생성 실패: {e}")

    return {"path": str(html_path), "size": html_path.stat().st_size,
            "url": _results_files_url(html_path)}


def _results_files_url(path: Path) -> str:
    """RESULTS_DIR 하위 파일 → 브라우저가 열 수 있는 /results-files/ URL.

    Linux 배포본은 결과가 `~/.local/share/ReplayKit/backend/results/...` (숨김 디렉토리)
    아래에 있는데, Ubuntu 기본 Firefox 는 snap 이라 `$HOME` 의 dot 디렉토리를 읽지 못한다
    (file:// 로 열면 "Access to the file was denied"). 리포트는 항상 실행 중인 서버를
    통해(HTTP) 열도록 URL 을 만들어 준다.
    """
    rel = path.resolve().relative_to(RESULTS_DIR.resolve()).as_posix()
    return "/results-files/" + "/".join(quote(seg) for seg in rel.split("/"))


def _result_html_path(filepath: Path) -> Path:
    """런 폴더면 폴더 내 result.html, 레거시 플랫 파일이면 같은 이름의 .html."""
    return (filepath.with_name("result.html")
            if filepath.name == "result.json"
            else filepath.with_suffix(".html"))


@router.get("/report/{filename:path}")
def open_result_report(filename: str):
    """result.html 로 리다이렉트 — 없으면 즉시 생성한다.

    프론트가 클릭 시점에 바로 새 탭으로 열 수 있게(비동기 대기 후 window.open 은
    팝업 차단됨) 리다이렉트로 처리한다.
    """
    try:
        filepath = (RESULTS_DIR / filename).resolve()
        filepath.relative_to(RESULTS_DIR.resolve())   # 경로 이탈 차단
    except (ValueError, OSError):
        raise HTTPException(status_code=400, detail="Invalid path")
    if not filepath.is_file():
        raise HTTPException(status_code=404, detail="Result not found")

    html_path = _result_html_path(filepath)
    if not html_path.is_file() or html_path.stat().st_size == 0:
        try:
            data = json.loads(filepath.read_text(encoding="utf-8"))
            html_path.write_text(_build_html_report(data, html_path), encoding="utf-8")
        except Exception as e:
            logger.exception("HTML report generation failed: %s", filename)
            raise HTTPException(status_code=500, detail=f"HTML 생성 실패: {e}")

    return RedirectResponse(url=_results_files_url(html_path), status_code=307)


@router.post("/open-folder")
async def open_result_folder(body: dict):
    """결과 파일이 있는 폴더를 파일 탐색기로 연다.

    filename 은 RESULTS_DIR 기준 상대경로. 파일이면 그 파일이 **선택된 상태**로
    부모 폴더를 열고(Windows: explorer /select), 폴더면 그 폴더를 연다.
    구간 저장 결과처럼 런 폴더 하위(`{run}/recordings/x.mp4`)도 바로 열 수 있다.
    """
    filename = body.get("filename", "")
    if not filename:
        raise HTTPException(status_code=400, detail="filename required")

    try:
        filepath = (RESULTS_DIR / filename).resolve()
        filepath.relative_to(RESULTS_DIR.resolve())   # 경로 이탈 차단
    except (ValueError, OSError):
        raise HTTPException(status_code=400, detail="Invalid path")
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")

    target = filepath if filepath.is_dir() else filepath.parent
    if sys.platform == "win32":
        if filepath.is_dir():
            os.startfile(str(target))
        else:
            # 파일을 선택된 상태로 열기 — 결과물이 어느 것인지 바로 보인다.
            subprocess.Popen(["explorer", f"/select,{filepath}"])
    else:
        subprocess.Popen(["xdg-open", str(target)])
    return {"status": "ok", "path": str(target)}


def _iter_run_dir_files(source_dir: Path):
    """런 폴더 내 파일을 순회. junction/symlink 디렉토리는 실제 대상을 따라감."""
    for item in sorted(source_dir.rglob("*")):
        if item.is_file():
            yield item


def _zip_directory(source_dir: Path, zip_path: Path, progress=None) -> None:
    """디렉토리를 ZIP 파일로 압축. progress(done, total): 파일 단위 진행 콜백(선택)."""
    files = list(_iter_run_dir_files(source_dir))
    total = len(files)
    with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
        for i, file in enumerate(files, start=1):
            arcname = file.relative_to(source_dir.parent).as_posix()
            zf.write(str(file), arcname)
            if progress is not None:
                try:
                    progress(i, total)
                except Exception:
                    pass


def _zip_directory_to_buffer(source_dir: Path, buf: io.BytesIO, progress=None) -> None:
    """디렉토리를 BytesIO 버퍼에 ZIP 압축. progress(done, total): 파일 단위 콜백(선택)."""
    files = list(_iter_run_dir_files(source_dir))
    total = len(files)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, file in enumerate(files, start=1):
            arcname = file.relative_to(source_dir.parent).as_posix()
            zf.write(str(file), arcname)
            if progress is not None:
                try:
                    progress(i, total)
                except Exception:
                    pass


@router.delete("/{filename:path}")
async def delete_result(filename: str):
    """Delete a test result and its associated files.

    런 폴더(folder/result.json) 또는 레거시 플랫 파일(.json) 모두 처리.
    """
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")

    deleted_recordings = []

    # 런 폴더인 경우 폴더 전체 삭제
    if filepath.name == "result.json" and filepath.parent != RESULTS_DIR:
        run_dir = filepath.parent
        folder_name = run_dir.name
        shutil.rmtree(str(run_dir), ignore_errors=True)
        # 연결된 웹캠 녹화 파일도 삭제 (webm + mp4)
        if RECORDINGS_DIR.is_dir():
            for pattern in (f"{folder_name}_webcam_*.webm", f"{folder_name}_webcam_*.mp4"):
                for rec in RECORDINGS_DIR.glob(pattern):
                    rec.unlink()
                    deleted_recordings.append(rec.name)
    else:
        filepath.unlink()
        base = filename.replace(".json", "")
        if RECORDINGS_DIR.is_dir():
            for pattern in (f"{base}_webcam_*.webm", f"{base}_webcam_*.mp4"):
                for rec in RECORDINGS_DIR.glob(pattern):
                    rec.unlink()
                    deleted_recordings.append(rec.name)

    return {"status": "deleted", "deleted_recordings": deleted_recordings}


def _safe_filename(name: str) -> str:
    """Path traversal 방어: 파일명에서 디렉토리 부분 제거."""
    return Path(name).name


# --- Webcam recording endpoints ---

@router.post("/webcam-upload")
async def upload_webcam_recording(
    file: UploadFile = File(...),
    result_filename: str = Query(...),
    repeat_index: int = Query(1),
    started_at: str | None = Query(None),
):
    """Upload a webcam recording linked to a test result.

    started_at: 브라우저 기반 녹화의 실제 시작 wall-clock 시각(ISO 문자열).
    사이드카(.meta.json)로 저장되어 프론트의 step→video 시간 매핑에 사용된다.
    """
    base = result_filename.replace(".json", "").replace("/result", "")
    filename = f"webcam_r{repeat_index}.webm"
    content = await file.read()

    # 시나리오 결과 폴더의 recordings/ 에 저장
    run_dir = RESULTS_DIR / base
    if run_dir.is_dir():
        rec_dir = run_dir / "recordings"
        rec_dir.mkdir(exist_ok=True)
        filepath = rec_dir / filename
        filepath.write_bytes(content)
    else:
        # 결과 폴더가 없으면 기존 위치에 저장 (폴백)
        RECORDINGS_DIR.mkdir(parents=True, exist_ok=True)
        filepath = RECORDINGS_DIR / f"{base}_webcam_r{repeat_index}.webm"
        filepath.write_bytes(content)

    # 사이드카에 시작 시각 저장 (있을 때만)
    if started_at:
        try:
            meta_path = filepath.with_suffix(filepath.suffix + ".meta.json")
            meta_path.write_text(
                json.dumps({"started_at": started_at}, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            pass

    return {"filename": filename, "path": str(filepath)}


def _read_recording_started_at(video_path: Path) -> str | None:
    """녹화 파일의 사이드카 (.meta.json)에서 started_at(wall-clock ISO)을 읽는다.

    파일이 없거나 파싱 실패 시 None. 프론트엔드는 None인 경우 첫 스텝 timestamp
    기반 휴리스틱으로 폴백한다.
    """
    try:
        meta_path = video_path.with_suffix(video_path.suffix + ".meta.json")
        if not meta_path.is_file():
            return None
        data = json.loads(meta_path.read_text(encoding="utf-8"))
        v = data.get("started_at")
        return v if isinstance(v, str) and v else None
    except Exception:
        return None


_CYCLE_RE = re.compile(r"(?:webcam|composite)_r(\d+)\.(?:webm|mp4)$")


def _recording_cycle_index(filename: str) -> int:
    """녹화 파일명에서 회차 번호를 추출. 매칭 실패 시 큰 값(맨 뒤 정렬)."""
    m = _CYCLE_RE.search(filename)
    return int(m.group(1)) if m else 10**9


@router.get("/recordings-for/{result_filename:path}")
async def list_recordings_for_result(result_filename: str):
    """List webcam recordings linked to a test result (both .webm and .mp4).

    - 회차 번호 기준 **숫자 정렬** (문자열 정렬이면 r1, r10, r11, r2... 로 뒤죽박죽 됨).
    - **0바이트 파일 제외** (카메라 끊김 등으로 생긴 재생불가 파일이 깨진 회차로 보이는 것 방지).
    - 같은 회차에 webm/mp4 가 둘 다 있으면 **하나만** 노출 (mp4 우선 → 더 큰 파일 우선).
    """
    RECORDINGS_DIR.mkdir(parents=True, exist_ok=True)
    base = result_filename.replace(".json", "").replace("/result", "")

    # 회차별 후보 수집 — key=cycle index, value=record dict
    candidates: dict[int, dict] = {}

    def _consider(f: Path, rel: str) -> None:
        try:
            size = f.stat().st_size
        except OSError:
            return
        if size == 0:
            return  # 빈/손상 파일 제외
        cycle = _recording_cycle_index(f.name)
        rec = {
            "filename": f.name,
            # 삭제/편집 API 가 실제 파일을 찾을 수 있는 경로. 파일명만으로는
            # 런 폴더 녹화를 특정할 수 없어 404 가 났었다.
            "rel_path": rel,
            "size": size,
            "url": f"/api/results/video/{quote(rel)}",
            "started_at": _read_recording_started_at(f),
            "_is_mp4": f.suffix.lower() == ".mp4",
        }
        prev = candidates.get(cycle)
        if prev is None:
            candidates[cycle] = rec
            return
        # 우선순위: mp4 > webm, 동급이면 큰 파일 우선
        better = (rec["_is_mp4"], rec["size"]) > (prev["_is_mp4"], prev["size"])
        if better:
            candidates[cycle] = rec

    # 런 폴더 내 recordings/ 확인 (webm + mp4)
    run_dir = RESULTS_DIR / base
    rec_dir = run_dir / "recordings" if run_dir.is_dir() else None
    if rec_dir and rec_dir.is_dir():
        for pattern in ("*.webm", "*.mp4"):
            for f in rec_dir.glob(pattern):
                # Range를 지원하는 전용 엔드포인트로 서빙 (StaticFiles는 206을 못 준다)
                _consider(f, f"{base}/recordings/{f.name}")

    # 레거시: Results/Video/ 에서도 탐색 (webm + mp4)
    for pattern in (f"{base}_webcam_*.webm", f"{base}_webcam_*.mp4"):
        for f in RECORDINGS_DIR.glob(pattern):
            _consider(f, f.name)

    recordings = [
        {k: v for k, v in rec.items() if k != "_is_mp4"}
        for _, rec in sorted(candidates.items(), key=lambda kv: kv[0])
    ]
    return {"recordings": recordings}


_RANGE_RE = re.compile(r"bytes=(\d*)-(\d*)")
_VIDEO_MIME = {".mp4": "video/mp4", ".webm": "video/webm", ".mkv": "video/x-matroska"}


def _resolve_recording(rel_path: str) -> Path | None:
    """녹화 상대경로 → 실제 파일. 런 폴더(results/) → 레거시(Results/Video/) 순.

    경로 이탈(../)은 relative_to 검사로 차단한다. 파일명만 주어지면 레거시에서만
    찾히므로, 런 폴더 파일은 프론트가 `{run}/recordings/{name}` 형태로 넘겨야 한다.
    """
    for root in (RESULTS_DIR, RECORDINGS_DIR):
        try:
            candidate = (root / rel_path).resolve()
            candidate.relative_to(root.resolve())
        except (ValueError, OSError):
            continue
        if candidate.is_file():
            return candidate
    return None


@router.get("/video/{rel_path:path}")
def stream_recording(rel_path: str, request: Request):
    """녹화 영상을 **Range 지원**으로 서빙한다.

    starlette 0.35의 StaticFiles는 Range를 처리하지 못해 206을 못 준다. 그러면
    브라우저가 `video.seekable`을 비워버려 seek이 0으로 snap되고, 프론트는 이를
    피하려고 **파일 전체를 blob으로 받아야만** 했다. 회차가 많은 에이징 결과에서는
    회차를 옮길 때마다 수십~수백 MB를 통째로 받느라 10초 seek 예산을 넘기고,
    그 다운로드들이 브라우저의 호스트당 6개 연결을 물고 있어 스크린샷 이미지까지
    같이 굶었다. 여기서 206을 제대로 주면 브라우저가 필요한 구간만 집어간다.

    sync def라 FastAPI가 스레드풀에서 실행 → 파일 IO가 이벤트 루프를 막지 않는다.
    """
    filepath = _resolve_recording(rel_path)
    if filepath is None:
        raise HTTPException(status_code=404, detail="Recording not found")

    media_type = _VIDEO_MIME.get(filepath.suffix.lower(), "application/octet-stream")
    file_size = filepath.stat().st_size
    base_headers = {"Accept-Ranges": "bytes", "Cache-Control": "no-cache"}

    range_header = request.headers.get("range")
    if not range_header:
        return FileResponse(str(filepath), media_type=media_type, headers=base_headers)

    m = _RANGE_RE.fullmatch(range_header.strip())
    if not m or file_size == 0:
        # 파싱 불가한 Range는 전체를 돌려준다(브라우저가 알아서 재시도).
        return FileResponse(str(filepath), media_type=media_type, headers=base_headers)

    start_s, end_s = m.groups()
    if start_s == "":
        # suffix range: 마지막 N 바이트 (moov atom이 뒤에 있는 mp4에서 실제로 쓰인다)
        if end_s == "":
            return FileResponse(str(filepath), media_type=media_type, headers=base_headers)
        start = max(0, file_size - int(end_s))
        end = file_size - 1
    else:
        start = int(start_s)
        end = int(end_s) if end_s else file_size - 1
    end = min(end, file_size - 1)

    if start >= file_size or start > end:
        return Response(status_code=416, headers={**base_headers,
                                                  "Content-Range": f"bytes */{file_size}"})

    length = end - start + 1

    def _iter_slice():
        remaining = length
        with open(filepath, "rb") as fh:
            fh.seek(start)
            while remaining > 0:
                chunk = fh.read(min(64 * 1024, remaining))
                if not chunk:
                    break
                remaining -= len(chunk)
                yield chunk

    return StreamingResponse(
        _iter_slice(),
        status_code=206,
        media_type=media_type,
        headers={
            **base_headers,
            "Content-Range": f"bytes {start}-{end}/{file_size}",
            "Content-Length": str(length),
        },
    )


@router.delete("/recordings/{filename:path}")
async def delete_recording(filename: str):
    """Delete a webcam recording (and its sidecar .meta.json if any).

    filename 은 런 폴더 기준 상대경로(`{run}/recordings/x.mp4`) 또는 레거시
    Results/Video 의 파일명. 예전엔 파일명만 받아 레거시 폴더에서만 찾았기 때문에
    런 폴더에 저장된 최신 녹화는 삭제/편집이 404 로 실패했다.
    """
    filepath = _resolve_recording(filename)
    if filepath is None:
        raise HTTPException(status_code=404, detail="Recording not found")
    safe_name = filepath.name
    filepath.unlink()
    # 사이드카 메타 파일도 함께 정리
    meta_path = filepath.with_suffix(filepath.suffix + ".meta.json")
    if meta_path.exists():
        try:
            meta_path.unlink()
        except Exception:
            pass
    return {"deleted": safe_name}


@router.post("/recordings/{filename:path}/trim")
def trim_recording(   # sync def: ffmpeg subprocess 가 블로킹이라 스레드풀에서 실행
    filename: str,
    start: float = Query(...),
    end: float = Query(...),
):
    """Trim a webcam recording (requires ffmpeg). 결과는 원본과 같은 폴더에 저장."""
    # 예상 못한 예외가 평문 "Internal Server Error"로 나가면 원인을 알 수 없다.
    # 여기서 트레이스백을 backend.log 에 남기고 예외 종류/메시지를 응답에 실어준다.
    try:
        return _trim_recording_impl(filename, start, end)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[trim] 예상 못한 오류 (filename=%r start=%s end=%s)",
                         filename, start, end)
        raise HTTPException(status_code=500,
                            detail=f"구간 저장 실패: {type(e).__name__}: {e}")


def _trim_recording_impl(filename: str, start: float, end: float) -> dict:
    filepath = _resolve_recording(filename)
    if filepath is None:
        raise HTTPException(status_code=404, detail="Recording not found")
    safe_name = filepath.name
    if start >= end:
        raise HTTPException(status_code=400, detail="start must be less than end")
    ffmpeg_path = _find_ffmpeg()
    if ffmpeg_path is None:
        raise HTTPException(
            status_code=400,
            detail="ffmpeg가 설치되어 있지 않습니다. tools/ 폴더에 ffmpeg.exe를 넣거나 시스템에 설치하세요."
        )
    output_name = f"trim_{start:.1f}_{end:.1f}_{safe_name}"
    # 원본과 같은 폴더에 저장 — 런 폴더 녹화를 레거시 폴더에 흘리지 않는다.
    output_path = filepath.parent / output_name

    # -ss/-to 는 반드시 **-i 앞**(입력 seek). 뒤에 두면(출력 seek) `-c copy` 와 겹쳐
    # 패킷이 잘못 버려져 2초를 요청해도 0.5초짜리 깨진 파일이 나온다(실측).
    # 또한 웹캠 녹화는 -g 를 지정하지 않아 키프레임 간격이 수 초~십수 초라
    # stream copy 로는 자를 위치가 크게 어긋난다 → 재인코딩으로 정확히 자른다.
    cmd = [
        ffmpeg_path,
        "-ss", str(start), "-to", str(end),
        "-i", str(filepath),
        "-c:v", "libx264", "-preset", "veryfast", "-crf", "23", "-pix_fmt", "yuv420p",
        "-c:a", "aac",              # 오디오가 없으면 무시된다
        "-movflags", "+faststart",
        str(output_path), "-y",
    ]
    try:
        subprocess.run(
            cmd, check=True, capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
        )
    except subprocess.CalledProcessError as e:
        stderr = (e.stderr or b"").decode(errors="replace").strip()
        # 전체를 로그로 남긴다 — 원인 규명은 보통 stderr 마지막 줄에 있다.
        logger.error("[trim] ffmpeg 실패 rc=%s\n  cmd: %s\n  stderr:\n%s",
                     e.returncode, " ".join(cmd), stderr)
        # 응답에는 **끝부분**을 싣는다. 앞 300자는 ffmpeg 버전 배너라 진단에 쓸모없다.
        tail = "\n".join(stderr.splitlines()[-6:])[-400:]
        raise HTTPException(status_code=500, detail=f"ffmpeg 실패 (rc={e.returncode}): {tail}")
    except OSError as e:
        logger.exception("[trim] ffmpeg 실행 불가: %s", ffmpeg_path)
        raise HTTPException(status_code=500, detail=f"ffmpeg 실행 실패: {e}")

    # rc=0 이어도 스트림이 없는 깨진 파일이 나올 수 있어 결과를 확인한다.
    if not output_path.exists() or output_path.stat().st_size == 0:
        logger.error("[trim] 결과 파일이 비어 있음: %s (cmd: %s)", output_path, " ".join(cmd))
        raise HTTPException(status_code=500, detail="구간 저장 결과 파일이 비어 있습니다.")

    # 서빙용 상대경로 — Range 지원 엔드포인트 기준
    for root in (RESULTS_DIR, RECORDINGS_DIR):
        try:
            rel = output_path.resolve().relative_to(root.resolve())
            break
        except ValueError:
            rel = None
    rel_str = str(rel).replace("\\", "/") if rel else output_name
    return {"filename": output_name, "rel_path": rel_str,
            # path: 사용자가 탐색기에서 바로 찾아갈 수 있는 절대 경로 (완료 알림에 표시)
            "path": str(output_path),
            "url": f"/api/results/video/{quote(rel_str)}"}


# update-step는 상세 모달의 BG_TASK 폴링이 스텝마다 호출한다. result.json이
# 에이징 결과처럼 수십~수백 MB면 read+dumps+write 한 번에 수 초가 걸리므로
# ① 스레드로 오프로드해 이벤트 루프를 막지 않고("서버 연결 중" 원인)
# ② 파일별 락으로 직렬화해 read-modify-write 유실을 막고
# ③ 락 대기 중 쌓인 요청을 한 번의 RMW로 합쳐 O(N²) 재직렬화를 없앤다.
_STEP_UPDATE_LOCKS: dict[str, asyncio.Lock] = {}
_STEP_UPDATE_PENDING: dict[str, list[dict]] = {}
_STEP_UPDATE_LAST_STATUS: dict[str, str] = {}


def _apply_step_updates(filepath: Path, updates: list[dict]) -> str:
    """큐에 쌓인 스텝 업데이트를 한 번의 읽기-쓰기로 모두 적용. (스레드 실행 전용)"""
    data = json.loads(filepath.read_text(encoding="utf-8"))
    step_results = data.get("step_results", [])
    status_map = {"pass": "passed_steps", "fail": "failed_steps",
                  "warning": "warning_steps", "error": "error_steps"}

    for body in updates:
        step_index = body.get("step_index")
        if step_index is None or step_index < 0 or step_index >= len(step_results):
            continue  # 이미 검증했지만 배치 중 하나가 어긋나도 나머지는 살린다
        sr = step_results[step_index]
        if "message" in body:
            sr["message"] = body["message"]
        if "status" in body:
            old_status = sr["status"]
            new_status = body["status"]
            sr["status"] = new_status
            # 카운트 재계산
            if old_status != new_status:
                if old_status in status_map:
                    data[status_map[old_status]] = max(0, data.get(status_map[old_status], 0) - 1)
                if new_status in status_map:
                    data[status_map[new_status]] = data.get(status_map[new_status], 0) + 1
                # 전체 상태 재평가
                if data.get("failed_steps", 0) > 0 or data.get("error_steps", 0) > 0:
                    data["status"] = "fail"
                elif data.get("warning_steps", 0) > 0:
                    data["status"] = "warning"
                else:
                    data["status"] = "pass"

    filepath.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return data.get("status", "")


@router.post("/update-step/{filename:path}")
async def update_step_result(filename: str, body: dict):
    """백그라운드 CMD 완료 후 스텝 결과를 영구 업데이트."""
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")
    step_index = body.get("step_index")
    if step_index is None or step_index < 0:
        raise HTTPException(status_code=400, detail="Invalid step_index")

    key = str(filepath)
    lock = _STEP_UPDATE_LOCKS.setdefault(key, asyncio.Lock())
    _STEP_UPDATE_PENDING.setdefault(key, []).append(body)

    async with lock:
        batch = _STEP_UPDATE_PENDING.pop(key, [])
        if not batch:
            # 앞선 요청이 내 업데이트까지 합쳐서 이미 기록함
            return {"status": "ok", "result_status": _STEP_UPDATE_LAST_STATUS.get(key, "")}
        result_status = await asyncio.to_thread(_apply_step_updates, filepath, batch)
        _STEP_UPDATE_LAST_STATUS[key] = result_status

    return {"status": "ok", "result_status": result_status}


@router.post("/update-steps/{filename:path}")
async def update_step_results_bulk(filename: str, body: dict):
    """여러 스텝 결과를 **한 번의 읽기-쓰기**로 업데이트.

    서버 재시작으로 BG_TASK 수백 개가 한꺼번에 소실되는 경우, 스텝마다
    update-step 을 부르면 대형 result.json 을 수백 번 재직렬화하게 된다
    (1회에 수 초 → 서버 전체가 밀려 영상 스트리밍까지 굶는다). 프론트가 모아서
    이 엔드포인트로 한 번에 보내면 쓰기는 1회로 끝난다.
    """
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")
    updates = body.get("updates")
    if not isinstance(updates, list) or not updates:
        raise HTTPException(status_code=400, detail="updates must be a non-empty list")
    for u in updates:
        si = u.get("step_index") if isinstance(u, dict) else None
        if si is None or si < 0:
            raise HTTPException(status_code=400, detail="Invalid step_index in updates")

    key = str(filepath)
    lock = _STEP_UPDATE_LOCKS.setdefault(key, asyncio.Lock())
    _STEP_UPDATE_PENDING.setdefault(key, []).extend(updates)

    async with lock:
        batch = _STEP_UPDATE_PENDING.pop(key, [])
        if not batch:
            return {"status": "ok", "applied": 0,
                    "result_status": _STEP_UPDATE_LAST_STATUS.get(key, "")}
        result_status = await asyncio.to_thread(_apply_step_updates, filepath, batch)
        _STEP_UPDATE_LAST_STATUS[key] = result_status

    return {"status": "ok", "applied": len(batch), "result_status": result_status}


@router.post("/migrate-legacy")
async def migrate_legacy():
    """레거시 결과 파일을 새 구조로 마이그레이션.
    screenshots/{name}/actual_{ts}/ → results/{ts}_{name}/screenshots/
    results/{name}_{ts}.json → results/{ts}_{name}/result.json
    """
    import re as _re
    migrated = 0
    errors = []

    # 1) screenshots 내 actual_ 폴더 → results 런 폴더로 이동
    if SCREENSHOTS_DIR.is_dir():
        for scenario_dir in SCREENSHOTS_DIR.iterdir():
            if not scenario_dir.is_dir():
                continue
            sc_name = scenario_dir.name
            for actual_dir in list(scenario_dir.iterdir()):
                if not actual_dir.is_dir() or not actual_dir.name.startswith("actual_"):
                    continue
                ts = actual_dir.name.replace("actual_", "")  # e.g. 20260408_174101
                if not _re.match(r"\d{8}_\d{6}", ts):
                    continue
                safe_name = _re.sub(r'[\\/:*?"<>|→]', '_', sc_name).replace(" ", "_")
                run_dir = RESULTS_DIR / f"{ts}_{safe_name}"
                run_dir.mkdir(parents=True, exist_ok=True)
                dst_ss = run_dir / "screenshots"
                if not dst_ss.exists():
                    try:
                        shutil.move(str(actual_dir), str(dst_ss))
                        migrated += 1
                    except Exception as e:
                        errors.append(f"screenshots/{sc_name}/{actual_dir.name}: {e}")
                else:
                    # 이미 존재하면 파일 단위로 머지
                    for f in actual_dir.iterdir():
                        if f.is_file():
                            dst_f = dst_ss / f.name
                            if not dst_f.exists():
                                shutil.move(str(f), str(dst_f))
                    # 빈 폴더 삭제
                    try:
                        actual_dir.rmdir()
                    except Exception:
                        pass
                    migrated += 1

    # 2) results 내 플랫 JSON → 런 폴더로 이동
    if RESULTS_DIR.is_dir():
        for json_file in list(RESULTS_DIR.glob("*.json")):
            # {name}_{timestamp}.json 패턴 매칭
            m = _re.match(r"^(.+?)_(\d{8}_\d{6})\.json$", json_file.name)
            if not m:
                continue
            sc_name = m.group(1)
            ts = m.group(2)
            safe_name = _re.sub(r'[\\/:*?"<>|→]', '_', sc_name).replace(" ", "_")
            run_dir = RESULTS_DIR / f"{ts}_{safe_name}"
            run_dir.mkdir(parents=True, exist_ok=True)
            dst = run_dir / "result.json"
            if not dst.exists():
                try:
                    shutil.move(str(json_file), str(dst))
                    # Excel도 함께 이동
                    xlsx = json_file.with_suffix(".xlsx")
                    if xlsx.exists():
                        shutil.move(str(xlsx), str(run_dir / "result.xlsx"))
                    migrated += 1
                except Exception as e:
                    errors.append(f"{json_file.name}: {e}")

    # 3) screenshots 내 actual/actual_ 폴더 정리 + 빈 폴더 삭제
    if SCREENSHOTS_DIR.is_dir():
        for d in list(SCREENSHOTS_DIR.iterdir()):
            if not d.is_dir():
                continue
            for sub in list(d.iterdir()):
                if sub.is_dir() and sub.name == "actual":
                    # 타임스탬프 없는 actual 폴더 (단일 스텝 테스트 임시) → 삭제
                    try:
                        shutil.rmtree(str(sub))
                        migrated += 1
                    except Exception as e:
                        errors.append(f"screenshots/{d.name}/actual: {e}")
            # 하위에 actual_ 폴더도 파일도 없으면 폴더 자체 삭제
            try:
                remaining = list(d.iterdir())
                if not remaining:
                    d.rmdir()
            except Exception:
                pass

    return {"migrated": migrated, "errors": errors}


@router.get("/{filename:path}")
async def get_result(filename: str):
    """Get a specific test result (런 폴더 또는 레거시 플랫 파일)."""
    filepath = RESULTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Result not found")
    # 에이징 결과는 수십 MB — 파싱/재직렬화 없이 원본 바이트를 그대로 흘려보낸다.
    # (dict로 반환하면 FastAPI가 이벤트 루프에서 다시 json.dumps 해 루프를 막는다)
    raw = await asyncio.to_thread(filepath.read_bytes)
    return Response(content=raw, media_type="application/json")


@router.get("/image/{scenario_name}/{image_path:path}")
async def get_image(scenario_name: str, image_path: str):
    """Serve a screenshot image."""
    filepath = SCREENSHOTS_DIR / scenario_name / image_path
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Image not found")
    return FileResponse(str(filepath), media_type="image/png")
